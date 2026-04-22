from tools.base import BaseTool, ToolResult
import os


PLUGIN_NAME = "excel"
PLUGIN_DESCRIPTION = "Read and write Excel/Google Sheets"

EXCEL_SCHEMA = {
    "type": "function",
    "function": {
        "name": "excel",
        "description": "Read, write, and analyze Excel and Google Sheets",
        "parameters": {
            "type": "object",
            "properties": {
                "action": {
                    "type": "string",
                    "enum": ["read", "write", "create_sheet", "list_sheets", "update_cell", "formula"],
                    "description": "Spreadsheet action"
                },
                "file": {"type": "string", "description": "Excel file path or sheet URL"},
                "sheet": {"type": "string", "description": "Sheet name"},
                "cell": {"type": "string", "description": "Cell (e.g., A1)"},
                "value": {"type": "string", "description": "Value to write"},
                "formula": {"type": "string", "description": "Formula to add"},
                "range": {"type": "string", "description": "Range (e.g., A1:C10)"},
            },
            "required": ["action"]
        }
    }
}


def run(action: str, file: str = "", sheet: str = "", cell: str = "", 
      value: str = "", formula: str = "", range: str = "") -> str:
    """Excel and Google Sheets operations."""
    try:
        import openpyxl
    except ImportError:
        return "Error: openpyxl not installed. Run: pip install openpyxl"
    
    try:
        # Load workbook
        if file.startswith("http"):
            # Google Sheets URL - would need service
            return "Error: Google Sheets requires service account setup"
        
        if file and os.path.exists(file):
            from openpyxl import load_workbook
            wb = load_workbook(file)
        else:
            return "Error: No file specified or file not found"
        
        if action == "list_sheets":
            sheets = wb.sheetnames
            return "Sheets:\n" + "\n".join([f"• {s}" for s in sheets])
        
        elif action == "read":
            ws = wb[sheet] if sheet else wb.active
            
            if range:
                data = []
                for row in ws[range]:
                    data.append([cell.value for cell in row])
                
                result = []
                for i, row in enumerate(data):
                    result.append(" | ".join([str(v)[:20] for v in row])))
                return "\n".join(result)
            
            # Read all
            data = []
            for row in ws.iter_rows(max_row=10):
                data.append([cell.value for cell in row])
            
            result = []
            for row in data:
                result.append(" | ".join([str(v)[:20] for v in row if v])))
            return "\n".join(result)
        
        elif action == "write" or action == "update_cell":
            ws = wb[sheet] if sheet else wb.active
            
            if cell:
                ws[cell] = value
            elif range:
                # Range write - value should be in CSV format
                values = value.split(",")
                row = ws[range.start.row]
                for i, v in enumerate(values):
                    if i < len(row):
                        row[i].value = v
            
            wb.save(file)
            return f"✓ Written to {file}"
        
        elif action == "create_sheet":
            ws = wb.create_sheet(title=sheet)
            wb.save(file)
            return f"✓ Created sheet: {sheet}"
        
        elif action == "formula":
            ws = wb[sheet] if sheet else wb.active
            ws[cell] = f"={formula}"
            wb.save(file)
            return f"✓ Added formula to {cell}"
        
        else:
            return f"Error: Unknown action {action}"
    
    except ImportError:
        return "Error: openpyxl not installed. Run: pip install openpyxl"
    except Exception as e:
        return f"Error: {str(e)}"


if __name__ == "__main__":
    print("Excel/Sheets plugin loaded.")